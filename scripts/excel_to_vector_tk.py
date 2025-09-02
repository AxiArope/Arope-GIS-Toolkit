import os, json
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText

import openpyxl
import shapefile  # 来自 pyshp 包
from pyproj import CRS

# =============== 基础工具 ===============

def read_excel_preview(path, sheet=None, max_rows=200):
    """读取 Excel 的表头与前 max_rows 行，用于预览与列名选择"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(sh.iter_rows(values_only=True))
    if not rows:
        wb.close()
        raise ValueError("Excel 文件为空")
    headers = [str(c) if c is not None else "" for c in rows[0]]
    data = rows[1:max_rows+1]
    wb.close()
    return wb.sheetnames, headers, data

def try_float(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except:
        return None

def infer_field_type(values, max_len=50):
    """非常简单的字段类型推断：全数值→F；全文本→C；混合→C"""
    has_text = False
    has_float = False
    for v in values[:200]:
        if v is None:
            continue
        if isinstance(v, (int, float)):
            has_float = True
        else:
            has_text = True
    if has_text and not has_float:
        return "C", max_len       # 文本
    if has_float and not has_text:
        return "F", 18            # 数值
    return "C", max_len           # 混合→文本

def write_shapefile(out_path, rows, headers, x_idx, y_idx, crs_epsg):
    """
    写出 ESRI Shapefile（点）。会生成 .shp/.shx/.dbf/.prj/.cpg
    rows: 数据行（不含表头），headers：表头列表
    x_idx/y_idx：X/Y 列索引
    """
    w = shapefile.Writer(out_path, shapeType=shapefile.POINT)

    # 除 X/Y 外的列都作为属性字段；SHP 字段名≤10字符
    attr_indices = [i for i in range(len(headers)) if i not in (x_idx, y_idx)]
    for i in attr_indices:
        name = (headers[i] or f"F{i}")[:10]
        col_vals = [r[i] if i is not None and i < len(r) else None for r in rows]
        ftype, flen = infer_field_type(col_vals)
        if ftype == "C":
            w.field(name, "C", size=flen)
        else:
            w.field(name, "F", size=18, decimal=6)

    count = 0
    for r in rows:
        x = try_float(r[x_idx] if x_idx < len(r) else None)
        y = try_float(r[y_idx] if y_idx < len(r) else None)
        if x is None or y is None:
            continue
        w.point(x, y)
        rec = []
        for i in attr_indices:
            v = r[i] if i < len(r) else None
            # DBF 不支持复杂类型，做个字符串化
            if isinstance(v, (list, dict, tuple, set)):
                v = str(v)
            rec.append(v)
        w.record(*rec)
        count += 1
    w.close()

    # 写 .prj（坐标系）
    crs = CRS.from_user_input(f"EPSG:{crs_epsg}")
    with open(os.path.splitext(out_path)[0] + ".prj", "w", encoding="utf-8") as f:
        f.write(crs.to_wkt())

    # 写 .cpg（编码声明，避免中文字段/值乱码）
    with open(os.path.splitext(out_path)[0] + ".cpg", "w", encoding="utf-8") as f:
        f.write("UTF-8")

    return count

def write_geojson(out_path, rows, headers, x_idx, y_idx, crs_epsg):
    """写出 GeoJSON（点要素集合）"""
    features = []
    for r in rows:
        x = try_float(r[x_idx] if x_idx < len(r) else None)
        y = try_float(r[y_idx] if y_idx < len(r) else None)
        if x is None or y is None:
            continue
        props = {}
        for i, h in enumerate(headers):
            if i in (x_idx, y_idx):
                continue
            v = r[i] if i < len(r) else None
            if isinstance(v, (list, dict, tuple, set)):
                v = str(v)
            props[h or f"F{i}"] = v
        feat = {
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [x, y]},
            "properties": props
        }
        features.append(feat)
    fc = {"type": "FeatureCollection", "features": features}
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(fc, f, ensure_ascii=False)
    return len(features)

# =============== Tk GUI ===============

class VectorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel → Shapefile / GeoJSON  by Arope")
        self.geometry("900x600")

        # 行 1：Excel 文件
        f1 = ttk.Frame(self); f1.pack(fill="x", padx=8, pady=6)
        ttk.Label(f1, text="Excel 文件").pack(side="left")
        self.var_file = tk.StringVar()
        ttk.Entry(f1, textvariable=self.var_file, width=70).pack(side="left", padx=6)
        ttk.Button(f1, text="浏览", command=self.pick_file).pack(side="left")

        # 行 2：Sheet
        f2 = ttk.Frame(self); f2.pack(fill="x", padx=8, pady=6)
        ttk.Label(f2, text="Sheet").pack(side="left")
        self.cmb_sheet = ttk.Combobox(f2, width=30, state="readonly")
        self.cmb_sheet.pack(side="left", padx=6)
        self.cmb_sheet.bind("<<ComboboxSelected>>", self.on_sheet_change)

        # 行 3：X/Y 列
        f3 = ttk.Frame(self); f3.pack(fill="x", padx=8, pady=6)
        ttk.Label(f3, text="X 列（经度/东向）").pack(side="left")
        self.cmb_x = ttk.Combobox(f3, width=27, state="readonly"); self.cmb_x.pack(side="left", padx=6)
        ttk.Label(f3, text="Y 列（纬度/北向）").pack(side="left")
        self.cmb_y = ttk.Combobox(f3, width=27, state="readonly"); self.cmb_y.pack(side="left", padx=6)

        # 行 4：坐标系 + 格式
        f4 = ttk.Frame(self); f4.pack(fill="x", padx=8, pady=6)
        ttk.Label(f4, text="坐标系 EPSG").pack(side="left")
        self.var_epsg = tk.StringVar(value="4326")
        ttk.Entry(f4, textvariable=self.var_epsg, width=12).pack(side="left", padx=6)
        self.var_fmt = tk.StringVar(value="shp")
        ttk.Radiobutton(f4, text="Shapefile (.shp)", variable=self.var_fmt, value="shp").pack(side="left", padx=12)
        ttk.Radiobutton(f4, text="GeoJSON (.geojson)", variable=self.var_fmt, value="geojson").pack(side="left")

        # 行 5：输出文件
        f5 = ttk.Frame(self); f5.pack(fill="x", padx=8, pady=6)
        ttk.Label(f5, text="输出文件").pack(side="left")
        self.var_out = tk.StringVar()
        ttk.Entry(f5, textvariable=self.var_out, width=70).pack(side="left", padx=6)
        ttk.Button(f5, text="选择", command=self.pick_save).pack(side="left")

        # 行 6：按钮
        f6 = ttk.Frame(self); f6.pack(fill="x", padx=8, pady=4)
        ttk.Button(f6, text="预览前 5 行", command=self.preview).pack(side="left")
        ttk.Button(f6, text="开始导出", command=self.run).pack(side="left", padx=8)
        ttk.Label(self, text="© 2025 by Arope", anchor="center").pack(side="bottom", pady=4)

        # 日志
        self.log = ScrolledText(self, height=16)
        self.log.pack(fill="both", expand=True, padx=8, pady=6)

        # 内部缓存
        self.headers = []
        self.preview_rows = []
        self.all_sheets = []

    # ---------- 日志 ----------
    def log_print(self, *msg):
        self.log.insert("end", " ".join(map(str, msg)) + "\n")
        self.log.see("end")

    # ---------- 事件 ----------
    def pick_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx;*.xls")])
        if not path:
            return
        self.var_file.set(path)
        try:
            sheets, headers, data = read_excel_preview(path)
            self.all_sheets = sheets
            self.headers = headers
            self.preview_rows = data

            self.cmb_sheet["values"] = sheets
            if sheets:
                self.cmb_sheet.set(sheets[0])

            self.cmb_x["values"] = headers
            self.cmb_y["values"] = headers

            self.log_print(f"已加载：{os.path.basename(path)}；sheets={sheets}；列={headers}")

            # 默认输出路径
            base, _ = os.path.splitext(path)
            default_ext = ".shp" if self.var_fmt.get() == "shp" else ".geojson"
            if not self.var_out.get():
                self.var_out.set(base + "_points" + default_ext)
        except Exception as e:
            messagebox.showerror("错误", f"读取 Excel 失败：{e}")

    def on_sheet_change(self, *_):
        path = self.var_file.get()
        sheet = self.cmb_sheet.get()
        if not (path and sheet):
            return
        try:
            _, headers, data = read_excel_preview(path, sheet=sheet, max_rows=200)
            self.headers = headers
            self.preview_rows = data
            self.cmb_x["values"] = headers
            self.cmb_y["values"] = headers
            self.log_print(f"切换到 sheet={sheet}；列={headers}")
        except Exception as e:
            messagebox.showerror("错误", f"读取 Sheet 失败：{e}")

    def preview(self):
        if not self.preview_rows:
            self.log_print("请先选择 Excel 文件。")
            return
        show = [self.headers] + self.preview_rows[:5]
        self.log_print("==== 预览 ====")
        for row in show:
            self.log_print(row)

    def pick_save(self):
        fmt = self.var_fmt.get()
        ext = ".shp" if fmt == "shp" else ".geojson"
        path = filedialog.asksaveasfilename(defaultextension=ext,
                                            filetypes=[("Shapefile", "*.shp"), ("GeoJSON", "*.geojson")])
        if path:
            self.var_out.set(path)

    # ---------- 导出 ----------
    def run(self):
        try:
            path = self.var_file.get()
            sheet = self.cmb_sheet.get()
            x_name = self.cmb_x.get()
            y_name = self.cmb_y.get()
            epsg = (self.var_epsg.get() or "4326").strip()
            out_path = self.var_out.get()
            fmt = self.var_fmt.get()

            if not (path and os.path.exists(path)):
                self.log_print("[错误] 请选择有效的 Excel 文件。"); return
            if not sheet:
                self.log_print("[错误] 请选择 Sheet。"); return
            if not (x_name and y_name):
                self.log_print("[错误] 请选择 X/Y 列。"); return
            if not out_path:
                self.log_print("[错误] 请选择输出文件。"); return
            if fmt not in ("shp", "geojson"):
                self.log_print("[错误] 只支持 shp 或 geojson。"); return

            # 读取整表
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sh = wb[sheet]
            rows = list(sh.iter_rows(values_only=True))
            wb_headers = [str(c) if c is not None else "" for c in rows[0]]
            if x_name not in wb_headers or y_name not in wb_headers:
                wb.close()
                self.log_print("[错误] X/Y 列名不在表头里。")
                return
            x_idx = wb_headers.index(x_name)
            y_idx = wb_headers.index(y_name)
            data_rows = rows[1:]
            wb.close()

            # 导出
            if fmt == "shp":
                n = write_shapefile(out_path, data_rows, wb_headers, x_idx, y_idx, epsg)
            else:
                n = write_geojson(out_path, data_rows, wb_headers, x_idx, y_idx, epsg)

            self.log_print(f"✅ 导出成功：{n} 个点 → {out_path}")
            messagebox.showinfo("成功", f"导出成功：{n} 个点\n{out_path}")
        except Exception as e:
            self.log_print("[错误]", e)
            messagebox.showerror("错误", str(e))

# =============== 入口 ===============
if __name__ == "__main__":
    VectorApp().mainloop()

